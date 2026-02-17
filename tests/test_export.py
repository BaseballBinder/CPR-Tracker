"""
Tests for the export service.
Tests PCO and Master template exports.
"""
import os
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.services.export_service import ExportService, get_export_service, ExportError
from app.models import SessionStatus


class TestExportService:
    """Tests for ExportService class."""

    def setup_method(self):
        """Set up test fixtures."""
        # Create a fresh ExportService instance (not the singleton)
        self.service = ExportService()

        # Create temp directories for templates and exports
        self.temp_dir = Path(tempfile.mkdtemp())
        self.template_dir = self.temp_dir / "templates_canroc"
        self.export_dir = self.temp_dir / "exports"
        self.template_dir.mkdir()
        self.export_dir.mkdir()

        # Save original settings for restoration
        self._orig_pco_path = self.service.settings.canroc_pco_template_path
        self._orig_master_path = self.service.settings.canroc_master_template_path
        self._orig_export_dir = self.service.settings.export_output_dir

        # Override settings paths for testing
        self.service.settings.canroc_pco_template_path = self.template_dir / "CanROC_PCO_Template.xlsx"
        self.service.settings.canroc_master_template_path = self.template_dir / "CanROC_Master_Template.xlsx"
        self.service.settings.export_output_dir = self.export_dir

    def teardown_method(self):
        """Clean up temp files and restore settings."""
        # Restore original settings
        self.service.settings.canroc_pco_template_path = self._orig_pco_path
        self.service.settings.canroc_master_template_path = self._orig_master_path
        self.service.settings.export_output_dir = self._orig_export_dir

        # Clean up temp directory
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def _create_pco_template(self, month_names: list = None):
        """Create a test PCO template with month tabs."""
        wb = Workbook()

        if month_names is None:
            month_names = self.service.MONTH_NAMES

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create month sheets with headers
        for month in month_names:
            ws = wb.create_sheet(title=month)

            # Add headers in Row 1
            headers = [
                "Date", "Time",
                "cr_cmprt1", "cr_cprff1", "cr_cdpth1", "cr_etco21", "cr_secun1",
                "cr_cmprt2", "cr_cprff2", "cr_cdpth2", "cr_etco22", "cr_secun2",
                "cr_cmprt3", "cr_cprff3", "cr_cdpth3", "cr_etco23", "cr_secun3",
            ]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

        wb.save(self.service.settings.canroc_pco_template_path)
        wb.close()

    def _create_master_template(self):
        """Create a test Master template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Add headers in Row 1
        headers = [
            "Date", "Time", "Event Type", "Outcome", "Provider",
            "Duration (s)", "Compression Rate", "Compression Depth",
            "CCF %", "Correct Depth %", "Correct Rate %"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(self.service.settings.canroc_master_template_path)
        wb.close()

    def _create_test_session(self, status: str = "complete", date: str = "2025-12-15"):
        """Create a test session dict."""
        return {
            "id": "test-session-001",
            "session_type": "real_call",
            "status": status,
            "date": date,
            "time": "14:30:00",
            "event_type": "Cardiac Arrest",
            "outcome": "ROSC",
            "provider_name": "John Smith",
            "metrics": {
                "duration": 480,
                "compression_rate": 112,
                "compression_depth": 5.4,
                "compression_fraction": 85,
                "correct_depth_percent": 78,
                "correct_rate_percent": 82,
                "cr_cmprt1": 108,
                "cr_cprff1": 82,
                "cr_cdpth1": 5.2,
                "cr_etco21": 35,
                "cr_secun1": 2,
                "cr_cmprt2": 112,
                "cr_cprff2": 88,
                "cr_cdpth2": 5.5,
                "cr_etco22": 38,
                "cr_secun2": 1,
                "cr_cmprt3": 115,
                "cr_cprff3": 90,
                "cr_cdpth3": 5.6,
                "cr_etco23": 40,
                "cr_secun3": 0,
            }
        }

    def test_export_pco_success(self):
        """Test successful PCO export."""
        self._create_pco_template()
        session = self._create_test_session()

        success, message, output_path = self.service.export_pco(session)

        assert success is True
        assert output_path is not None
        assert output_path.exists()
        assert "CanROC_PCO" in output_path.name

    def test_export_pco_selects_correct_month(self):
        """Test that PCO export selects the correct month tab."""
        self._create_pco_template()
        session = self._create_test_session(date="2025-03-15")  # March

        success, message, output_path = self.service.export_pco(session)

        assert success is True

        # Load the exported file and verify March tab was used
        wb = load_workbook(output_path)
        ws = wb["March"]

        # Check that data was written to row 2
        assert ws.cell(row=2, column=1).value == "2025-03-15"
        wb.close()

    def test_export_pco_writes_correct_values(self):
        """Test that PCO export writes correct values to correct columns."""
        self._create_pco_template()
        session = self._create_test_session(date="2025-12-15")

        success, message, output_path = self.service.export_pco(session)

        assert success is True

        # Load the exported file and verify values
        wb = load_workbook(output_path)
        ws = wb["December"]

        # Check specific values (based on header positions in our test template)
        assert ws.cell(row=2, column=1).value == "2025-12-15"  # Date
        assert ws.cell(row=2, column=2).value == "14:30:00"  # Time
        assert ws.cell(row=2, column=3).value == 108  # cr_cmprt1
        assert ws.cell(row=2, column=4).value == 82  # cr_cprff1
        assert ws.cell(row=2, column=5).value == 5.2  # cr_cdpth1

        wb.close()

    def test_export_pco_preserves_headers(self):
        """Test that PCO export preserves original headers."""
        self._create_pco_template()
        session = self._create_test_session()

        success, message, output_path = self.service.export_pco(session)

        assert success is True

        # Load original template and exported file
        wb_original = load_workbook(self.service.settings.canroc_pco_template_path)
        wb_exported = load_workbook(output_path)

        ws_orig = wb_original["December"]
        ws_exp = wb_exported["December"]

        # Compare headers in Row 1
        for col in range(1, 20):
            original_header = ws_orig.cell(row=1, column=col).value
            exported_header = ws_exp.cell(row=1, column=col).value
            assert original_header == exported_header, f"Header mismatch at column {col}"

        wb_original.close()
        wb_exported.close()

    def test_export_pco_appends_row(self):
        """Test that PCO export appends rows (doesn't overwrite)."""
        self._create_pco_template()

        # Export first session
        session1 = self._create_test_session(date="2025-12-10")
        session1["id"] = "session-001"
        success1, _, output_path1 = self.service.export_pco(session1)

        assert success1 is True

        # Copy the exported file back to template location to simulate existing data
        shutil.copy(output_path1, self.service.settings.canroc_pco_template_path)

        # Export second session
        session2 = self._create_test_session(date="2025-12-15")
        session2["id"] = "session-002"
        success2, _, output_path2 = self.service.export_pco(session2)

        assert success2 is True

        # Verify both rows exist
        wb = load_workbook(output_path2)
        ws = wb["December"]

        assert ws.cell(row=2, column=1).value == "2025-12-10"  # First session
        assert ws.cell(row=3, column=1).value == "2025-12-15"  # Second session

        wb.close()

    def test_export_pco_incomplete_session_fails(self):
        """Test that PCO export fails for incomplete sessions."""
        self._create_pco_template()
        session = self._create_test_session(status="importing")

        success, message, output_path = self.service.export_pco(session)

        assert success is False
        assert "importing" in message.lower()
        assert output_path is None

    def test_export_pco_missing_template_fails(self):
        """Test that PCO export fails when template is missing."""
        # Don't create template
        session = self._create_test_session()

        success, message, output_path = self.service.export_pco(session)

        assert success is False
        assert "not found" in message.lower()

    def test_export_master_success(self):
        """Test successful Master export."""
        self._create_master_template()
        session = self._create_test_session()

        success, message, output_path = self.service.export_master(session)

        assert success is True
        assert output_path is not None
        assert output_path.exists()
        assert "CanROC_Master" in output_path.name

    def test_export_master_writes_correct_values(self):
        """Test that Master export writes correct values."""
        self._create_master_template()
        session = self._create_test_session()

        success, message, output_path = self.service.export_master(session)

        assert success is True

        # Load and verify
        wb = load_workbook(output_path)
        ws = wb["Data"]

        # Check values (based on our test template header order)
        assert ws.cell(row=2, column=1).value == "2025-12-15"  # Date
        assert ws.cell(row=2, column=2).value == "14:30:00"  # Time
        assert ws.cell(row=2, column=3).value == "Cardiac Arrest"  # Event Type
        assert ws.cell(row=2, column=4).value == "ROSC"  # Outcome
        assert ws.cell(row=2, column=5).value == "John Smith"  # Provider
        assert ws.cell(row=2, column=6).value == 480  # Duration

        wb.close()

    def test_export_master_preserves_headers(self):
        """Test that Master export preserves original headers."""
        self._create_master_template()
        session = self._create_test_session()

        success, message, output_path = self.service.export_master(session)

        assert success is True

        # Load original template and exported file
        wb_original = load_workbook(self.service.settings.canroc_master_template_path)
        wb_exported = load_workbook(output_path)

        ws_orig = wb_original["Data"]
        ws_exp = wb_exported["Data"]

        # Compare headers in Row 1
        for col in range(1, 12):
            original_header = ws_orig.cell(row=1, column=col).value
            exported_header = ws_exp.cell(row=1, column=col).value
            assert original_header == exported_header, f"Header mismatch at column {col}"

        wb_original.close()
        wb_exported.close()

    def test_get_month_from_date(self):
        """Test month extraction from date string."""
        assert self.service._get_month_from_date("2025-01-15") == "January"
        assert self.service._get_month_from_date("2025-06-01") == "June"
        assert self.service._get_month_from_date("2025-12-31") == "December"

    def test_get_month_from_invalid_date(self):
        """Test month extraction from invalid date returns current month."""
        result = self.service._get_month_from_date("invalid")
        assert result in self.service.MONTH_NAMES

    def test_get_available_templates(self):
        """Test checking template availability."""
        # Before creating templates
        templates = self.service.get_available_templates()
        assert templates["pco"] is False
        assert templates["master"] is False

        # After creating templates
        self._create_pco_template()
        self._create_master_template()

        templates = self.service.get_available_templates()
        assert templates["pco"] is True
        assert templates["master"] is True

    def test_deterministic_output_filename(self):
        """Test that output filenames follow CanROC format."""
        self._create_pco_template()
        session = self._create_test_session()

        filename1 = self.service._generate_output_filename(session, "pco")
        filename2 = self.service._generate_output_filename(session, "master")

        # Should follow CanROC_<TYPE>_<date>_<session_id>.xlsx format
        assert filename1 == "CanROC_PCO_2025-12-15_test-ses.xlsx"
        assert filename2 == "CanROC_Master_2025-12-15_test-ses.xlsx"

        # Should contain date
        assert "2025-12-15" in filename1
        assert "2025-12-15" in filename2

        # Should contain session ID prefix
        assert "test-ses" in filename1


    def test_export_pco_validates_headers(self):
        """Test that PCO export validates required headers."""
        # Create template without Date and Time columns
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for month in self.service.MONTH_NAMES:
            ws = wb.create_sheet(title=month)
            # Only add PCO metric columns, not Date/Time
            headers = ["cr_cmprt1", "cr_cprff1", "cr_cdpth1"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

        wb.save(self.service.settings.canroc_pco_template_path)
        wb.close()

        session = self._create_test_session()

        success, message, output_path = self.service.export_pco(session)

        assert success is False
        assert "Missing required headers" in message
        assert "Date" in message or "Time" in message

    def test_export_master_validates_headers(self):
        """Test that Master export validates required headers."""
        # Create template without Date column
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Only add non-Date columns
        headers = ["Time", "Event Type", "Outcome"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(self.service.settings.canroc_master_template_path)
        wb.close()

        session = self._create_test_session()

        success, message, output_path = self.service.export_master(session)

        assert success is False
        assert "Missing required headers" in message
        assert "Date" in message


class TestExportServiceSingleton:
    """Test singleton behavior."""

    def test_get_export_service_returns_same_instance(self):
        """Test that get_export_service returns the same instance."""
        service1 = get_export_service()
        service2 = get_export_service()
        assert service1 is service2
