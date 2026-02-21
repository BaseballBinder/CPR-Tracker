"""
Export service for CanROC template exports.
Handles PCO and Master template population and export.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, Tuple

logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import get_settings
from app.models import SessionStatus


class ExportError(Exception):
    """Exception raised when export fails."""
    pass


def ensure_templates():
    """
    Verify CanROC Excel template files exist.
    Called on application startup to check templates are available.

    Note: We use the original CanROC templates uploaded by the user.
    These templates should NOT be auto-generated as they have specific
    CanROC field structures and existing data.
    """
    settings = get_settings()

    # Check PCO template
    if not settings.canroc_pco_template_path.exists():
        logger.warning(f"PCO template not found at {settings.canroc_pco_template_path}")

    # Check Master template
    if not settings.canroc_master_template_path.exists():
        logger.warning(f"Master template not found at {settings.canroc_master_template_path}")


class ExportService:
    """Service for exporting session data to CanROC templates."""

    # Month tab names as they appear in the actual PCO template
    # Note: Some have trailing spaces in the original file
    MONTH_TAB_NAMES = [
        "Jan ", "Feb ", "Mar ", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]

    # PCO column headers - these are the exact headers from the original template
    # PCO minute data: cr_cmprt, cr_cprff, cr_cdpth, cr_etco2, cr_secun (x10 minutes)
    # Plus additional fields for shock times, duration, etc.
    PCO_FIELDS = [
        # Per-minute CPR metrics (minutes 1-10)
        *[f"cr_cmprt{i}" for i in range(1, 11)],   # Compression rate
        *[f"cr_cprff{i}" for i in range(1, 11)],   # CPR fraction
        *[f"cr_cdpth{i}" for i in range(1, 11)],   # Compression depth
        *[f"cr_etco2{i}" for i in range(1, 11)],   # End-tidal CO2
        *[f"cr_secun{i}" for i in range(1, 11)],   # Seconds uncompressed
        # Additional PCO fields
        "cr_cprprc",      # CPR process
        "cr_duration",    # Duration
        "cr_pearate",     # PEA rate
        "cr_cmpfrag",     # Compression fragment
        "cr_comprtag",    # Compression rate aggregate
        "cr_ventrate",    # Ventilation rate
        "cr_numpau>10s",  # Number of pauses >10s
        "cr_longpause",   # Longest pause
        "cr_timrosc",     # Time to ROSC
        "cr_etco2",       # End-tidal CO2 (aggregate)
    ]

    # Master template column headers - CanROC field names from the original template
    # The Master sheet has 148 columns - we only populate the ones relevant to our data
    MASTER_FIELDS = [
        "pcofile",        # PCO file reference (col 1)
        "cr_epdt",        # Event/Patient date (col 10)
        "cr_service ",    # Service (col 9, note trailing space)
        "cr_duration",    # Duration
    ]

    def __init__(self):
        self.settings = get_settings()

    def export_pco(self, session: Dict[str, Any]) -> Tuple[bool, str, Optional[Path]]:
        """
        Export session data to PCO template.

        Args:
            session: Session data dict with metrics and canroc_pco_payload

        Returns:
            Tuple of (success, message, output_path)
        """
        # Validate session status
        if session.get("status") != SessionStatus.COMPLETE.value:
            return False, f"Cannot export: session status is '{session.get('status')}', not 'complete'", None

        # Check template exists
        template_path = self.settings.canroc_pco_template_path
        if not template_path.exists():
            return False, f"PCO template not found: {template_path}", None

        try:
            # Load template
            wb = load_workbook(template_path)

            # Get month tab from session date
            session_date = session.get("date", "")
            month_tab = self._get_month_tab_from_date(session_date)

            # Find the correct month tab
            if month_tab not in wb.sheetnames:
                # Try without trailing space
                month_tab_stripped = month_tab.strip()
                if month_tab_stripped in wb.sheetnames:
                    month_tab = month_tab_stripped
                else:
                    return False, f"Month tab '{month_tab}' not found in PCO template. Available: {wb.sheetnames}", None

            ws = wb[month_tab]

            # Build header -> column index mapping from Row 1
            header_map = self._build_header_map(ws)

            # Validate headers before proceeding
            valid, error_msg = self._validate_pco_headers(header_map)
            if not valid:
                wb.close()
                return False, error_msg, None

            # Build payload from session
            payload = self._build_pco_payload(session)

            # Find the next available row for data entry
            next_row = self._find_next_available_row(ws, start_row=4, check_col=header_map.get("cr_cmprt1", 7))

            # Write values to the row - match payload fields to template headers
            for field, value in payload.items():
                if field in header_map and value is not None:
                    col_idx = header_map[field]
                    ws.cell(row=next_row, column=col_idx, value=value)

            # Generate output filename
            output_filename = self._generate_output_filename(session, "pco")
            output_path = self.settings.export_output_dir / output_filename

            # Save the workbook
            wb.save(output_path)
            wb.close()

            return True, f"PCO export saved to {output_filename}", output_path

        except Exception as e:
            return False, f"Export failed: {e}", None

    def export_master(self, session: Dict[str, Any]) -> Tuple[bool, str, Optional[Path]]:
        """
        Export session data to Master template.

        Args:
            session: Session data dict with metrics

        Returns:
            Tuple of (success, message, output_path)
        """
        # Validate session status
        if session.get("status") != SessionStatus.COMPLETE.value:
            return False, f"Cannot export: session status is '{session.get('status')}', not 'complete'", None

        # Check template exists
        template_path = self.settings.canroc_master_template_path
        if not template_path.exists():
            return False, f"Master template not found: {template_path}", None

        try:
            # Load template
            wb = load_workbook(template_path)

            # Use "Master" sheet (the main data sheet in the original template)
            if "Master" in wb.sheetnames:
                ws = wb["Master"]
            else:
                ws = wb.active

            # Build header -> column index mapping from Row 1
            header_map = self._build_header_map(ws)

            # Validate headers before proceeding
            valid, error_msg = self._validate_master_headers(header_map)
            if not valid:
                wb.close()
                return False, error_msg, None

            # Build payload from session
            payload = self._build_master_payload(session)

            # Find the next available row for data entry
            # Check column 1 (pcofile) for Master template
            next_row = self._find_next_available_row(ws, start_row=4, check_col=1)

            # Write values to the row - match payload fields to template headers
            for field, value in payload.items():
                if field in header_map and value is not None:
                    col_idx = header_map[field]
                    ws.cell(row=next_row, column=col_idx, value=value)

            # Generate output filename
            output_filename = self._generate_output_filename(session, "master")
            output_path = self.settings.export_output_dir / output_filename

            # Save the workbook
            wb.save(output_path)
            wb.close()

            return True, f"Master export saved to {output_filename}", output_path

        except Exception as e:
            return False, f"Export failed: {e}", None

    def _get_month_tab_from_date(self, date_str: str) -> str:
        """Get month tab name from date string (YYYY-MM-DD format)."""
        try:
            if not date_str:
                # Default to current month
                return self.MONTH_TAB_NAMES[datetime.now().month - 1]

            parts = date_str.split("-")
            if len(parts) >= 2:
                month_num = int(parts[1])
                if 1 <= month_num <= 12:
                    return self.MONTH_TAB_NAMES[month_num - 1]

            return self.MONTH_TAB_NAMES[datetime.now().month - 1]
        except (ValueError, IndexError):
            return self.MONTH_TAB_NAMES[datetime.now().month - 1]

    def _build_header_map(self, ws: Worksheet) -> Dict[str, int]:
        """
        Build a mapping of header names to column indices.
        Reads Row 1 of the worksheet.
        """
        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header_map[str(cell.value).strip()] = col_idx
        return header_map

    def _validate_pco_headers(self, header_map: Dict[str, int]) -> Tuple[bool, str]:
        """
        Validate that required PCO headers exist in the template.
        Returns (valid, error_message).
        """
        # Check for at least some PCO minute fields (cr_cmprt1 is a good indicator)
        if "cr_cmprt1" not in header_map:
            return False, "Missing PCO minute data headers (cr_cmprt1 not found)"

        return True, ""

    def _validate_master_headers(self, header_map: Dict[str, int]) -> Tuple[bool, str]:
        """
        Validate that required Master headers exist in the template.
        Returns (valid, error_message).
        """
        # Check for pcofile column (first column in the Master template)
        if "pcofile" not in header_map:
            return False, "Missing 'pcofile' header in Master template"

        return True, ""

    def _find_next_available_row(self, ws: Worksheet, start_row: int = 4, check_col: int = 7) -> int:
        """
        Find the next row available for data entry.

        Scans from start_row looking for the first row where the check column
        does NOT contain real data (meaning it has None, '', or '.' placeholder).
        Then continues to find where real data ends and appends after that.

        Args:
            ws: The worksheet to search
            start_row: Row to start searching from (default 4 to skip header rows)
            check_col: Column to check for data (default 7, which is typically a data column)

        Returns:
            The row number where new data should be written
        """
        row = start_row
        last_data_row = start_row - 1  # Track the last row with real data

        while row < 10000:  # Safety limit
            val = ws.cell(row=row, column=check_col).value

            # Check if this row has real data (not None, '', or '.' placeholder)
            is_real_data = val is not None and str(val).strip() not in ('', '.')

            if is_real_data:
                last_data_row = row
            elif val is None:
                # Hit a completely empty row - append here if we've seen data
                # or keep going if we haven't found any data yet
                break

            row += 1

        # Return the row after the last data row
        return last_data_row + 1

    def _build_pco_payload(self, session: Dict[str, Any]) -> Dict[str, Any]:
        """Build PCO payload from session data."""
        payload = {
            "date": session.get("date", ""),
            "time": session.get("time", ""),
        }

        # Get metrics
        metrics = session.get("metrics", {})
        if isinstance(metrics, dict):
            # Add PCO minute fields
            for minute in range(1, 11):
                payload[f"cr_cmprt{minute}"] = metrics.get(f"cr_cmprt{minute}")
                payload[f"cr_cprff{minute}"] = metrics.get(f"cr_cprff{minute}")
                payload[f"cr_cdpth{minute}"] = metrics.get(f"cr_cdpth{minute}")
                payload[f"cr_etco2{minute}"] = metrics.get(f"cr_etco2{minute}")
                payload[f"cr_secun{minute}"] = metrics.get(f"cr_secun{minute}")

        # Also check canroc_pco_payload if present
        pco_payload = session.get("canroc_pco_payload", {})
        if isinstance(pco_payload, dict):
            payload.update(pco_payload)

        return payload

    def _build_master_payload(self, session: Dict[str, Any]) -> Dict[str, Any]:
        """
        Build Master payload from session data.
        Uses CanROC field names that match the Master template headers.
        """
        payload = {}

        # Map session data to CanROC Master fields
        # pcofile: Reference to PCO file (use session ID)
        payload["pcofile"] = session.get("id", "")

        # cr_epdt: Event/patient date
        payload["cr_epdt"] = session.get("date", "")

        # Get metrics and map to CanROC fields
        metrics = session.get("metrics", {})
        if isinstance(metrics, dict):
            # cr_duration: Duration in seconds
            if "duration" in metrics:
                payload["cr_duration"] = metrics.get("duration")

            # Include any canroc-prefixed fields from metrics directly
            for key, value in metrics.items():
                if key.startswith("cr_") and value is not None:
                    payload[key] = value

        # Also check for canroc_master_payload if present (for direct field mapping)
        master_payload = session.get("canroc_master_payload", {})
        if isinstance(master_payload, dict):
            payload.update(master_payload)

        return payload

    def _generate_output_filename(self, session: Dict[str, Any], export_type: str) -> str:
        """
        Generate deterministic output filename.

        Format: CanROC_<TYPE>_<date>_<session_id>.xlsx
        Examples:
            - CanROC_PCO_2025-12-30_abc12345.xlsx
            - CanROC_Master_2025-12-30_abc12345.xlsx
        """
        session_id = session.get("id", "unknown")
        date = session.get("date", datetime.now().strftime("%Y-%m-%d"))

        # Map export_type to proper case
        type_name = "PCO" if export_type == "pco" else "Master"

        return f"CanROC_{type_name}_{date}_{session_id[:8]}.xlsx"

    def get_available_templates(self) -> Dict[str, bool]:
        """Check which templates are available."""
        return {
            "pco": self.settings.canroc_pco_template_path.exists(),
            "master": self.settings.canroc_master_template_path.exists(),
        }


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get the singleton export service instance."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
