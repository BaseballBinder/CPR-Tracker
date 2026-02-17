"""
Schema Service for CanROC Wizard.
Handles loading, validation, and template drift detection for CanROC schemas.
"""
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from functools import lru_cache

import openpyxl

logger = logging.getLogger(__name__)

from app.desktop_config import get_bundle_dir
from app.service_context import get_active_service_dir

def _get_schemas_dir():
    """Get schemas directory - from active service if available, else bundle."""
    service_dir = get_active_service_dir()
    if service_dir:
        return service_dir / "data" / "schemas"
    return get_bundle_dir() / "data" / "schemas"

def _get_templates_dir():
    """Get CanROC templates directory."""
    service_dir = get_active_service_dir()
    if service_dir:
        return service_dir / "templates_canroc"
    return get_bundle_dir() / "templates_canroc"

SCHEMA_FILES = {
    "master": lambda: _get_schemas_dir() / "canroc_master_schema.json",
    "pco": lambda: _get_schemas_dir() / "canroc_pco_schema.json",
}

TEMPLATE_FILES = {
    "master": lambda: _get_templates_dir() / "1.Master_CanROC_Sheet_Update_August 2025.xlsx",
    "pco": lambda: _get_templates_dir() / "4. CanROC_Variables_PCO_Files_Master_Update_June2025.xlsx",
}


class SchemaService:
    """Service for managing CanROC schemas."""

    def __init__(self):
        self._schemas: Dict[str, Dict] = {}
        self._field_index: Dict[str, Dict[str, Dict]] = {}  # template_id -> field_id -> field_def

    def load_schema(self, template_id: str) -> Dict[str, Any]:
        """
        Load schema for a template.

        Args:
            template_id: "master" or "pco"

        Returns:
            Full schema dictionary

        Raises:
            FileNotFoundError: If schema file doesn't exist
            ValueError: If template_id is invalid
        """
        if template_id not in SCHEMA_FILES:
            raise ValueError(f"Invalid template_id: {template_id}. Must be 'master' or 'pco'")

        if template_id in self._schemas:
            return self._schemas[template_id]

        schema_path = SCHEMA_FILES[template_id]()
        if not schema_path.exists():
            raise FileNotFoundError(f"Schema file not found: {schema_path}")

        with open(schema_path, "r", encoding="utf-8") as f:
            schema = json.load(f)

        self._schemas[template_id] = schema
        self._build_field_index(template_id, schema)

        return schema

    def _build_field_index(self, template_id: str, schema: Dict) -> None:
        """Build a field_id -> field_def index for fast lookup."""
        index = {}
        for page in schema.get("pages", []):
            for field in page.get("fields", []):
                field_id = field.get("field_id")
                if field_id:
                    # Add page context to field
                    field_with_context = {**field, "page_id": page["page_id"], "page_name": page["page_name"]}
                    index[field_id] = field_with_context
        self._field_index[template_id] = index

    def get_field(self, template_id: str, field_id: str) -> Optional[Dict]:
        """
        Get a field definition by ID.

        Args:
            template_id: "master" or "pco"
            field_id: The cr_* field code

        Returns:
            Field definition dict or None if not found
        """
        if template_id not in self._field_index:
            self.load_schema(template_id)

        return self._field_index.get(template_id, {}).get(field_id)

    def get_all_field_ids(self, template_id: str) -> List[str]:
        """Get all field IDs for a template."""
        if template_id not in self._field_index:
            self.load_schema(template_id)

        return list(self._field_index.get(template_id, {}).keys())

    def get_page(self, template_id: str, page_id: int) -> Optional[Dict]:
        """
        Get a page definition by ID.

        Args:
            template_id: "master" or "pco"
            page_id: 1-indexed page number

        Returns:
            Page definition dict or None if not found
        """
        schema = self.load_schema(template_id)
        for page in schema.get("pages", []):
            if page.get("page_id") == page_id:
                return page
        return None

    def get_total_pages(self, template_id: str) -> int:
        """Get total number of pages in a template."""
        schema = self.load_schema(template_id)
        return len(schema.get("pages", []))

    def get_required_fields(self, template_id: str) -> List[str]:
        """Get list of required field IDs for a template."""
        schema = self.load_schema(template_id)
        completion_rules = schema.get("completion_rules", {})
        return completion_rules.get("required_fields", [])

    def get_missing_marker(self, template_id: str) -> str:
        """Get the missing marker character for a template (default '.')."""
        schema = self.load_schema(template_id)
        return schema.get("missing_marker", ".")

    def validate_schema_against_template(self, template_id: str) -> List[str]:
        """
        Validate schema field_ids match Excel template Row 1 codes.

        Args:
            template_id: "master" or "pco"

        Returns:
            List of warnings/errors. Empty list means validation passed.
        """
        warnings = []

        template_fn = TEMPLATE_FILES.get(template_id)
        template_path = template_fn() if template_fn else None
        if not template_path or not template_path.exists():
            warnings.append(f"CRITICAL: Template file not found: {template_path}")
            return warnings

        try:
            schema = self.load_schema(template_id)
        except Exception as e:
            warnings.append(f"CRITICAL: Failed to load schema: {e}")
            return warnings

        try:
            wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)

            # Determine which sheet to use
            if template_id == "master":
                ws = wb["Master"]
            else:
                # For PCO, use any month sheet (they should have same structure)
                ws = wb["Jan "]  # Note the trailing space

            # Extract Row 1 codes from Excel
            excel_codes = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    code = str(cell_value).strip()
                    excel_codes[code] = col

            wb.close()

            # Get schema field IDs
            schema_field_ids = self.get_all_field_ids(template_id)

            # Check for missing fields in Excel
            for field_id in schema_field_ids:
                if field_id not in excel_codes:
                    warnings.append(f"MISSING: Field '{field_id}' in schema but not in Excel Row 1")

            # Check for fields in Excel not in schema
            for code in excel_codes:
                if code not in schema_field_ids:
                    # Only warn about cr_* fields
                    if code.startswith("cr_") or code in ["pcofile", "ptid", "ptid2"]:
                        warnings.append(f"EXTRA: Field '{code}' in Excel but not in schema")

            # Verify column positions match
            for field_id in schema_field_ids:
                field_def = self.get_field(template_id, field_id)
                if field_def and field_id in excel_codes:
                    expected_col = field_def.get("excel_column")
                    actual_col = excel_codes[field_id]
                    if expected_col and expected_col != actual_col:
                        warnings.append(
                            f"MOVED: Field '{field_id}' expected column {expected_col}, found at {actual_col}"
                        )

        except Exception as e:
            warnings.append(f"ERROR: Failed to validate against template: {e}")

        return warnings

    def validate_all_schemas(self) -> Dict[str, List[str]]:
        """
        Validate all schemas against their templates.

        Returns:
            Dict mapping template_id to list of warnings
        """
        results = {}
        for template_id in SCHEMA_FILES:
            results[template_id] = self.validate_schema_against_template(template_id)
        return results

    def get_field_choices(self, template_id: str, field_id: str) -> Optional[List[Dict]]:
        """Get choices for a field if it's a choice type."""
        field = self.get_field(template_id, field_id)
        if field and field.get("type") == "choice":
            return field.get("choices", [])
        return None

    def get_field_dependencies(self, template_id: str, field_id: str) -> List[Dict]:
        """Get dependencies for a field."""
        field = self.get_field(template_id, field_id)
        if field:
            return field.get("dependencies", [])
        return []

    def is_cno_allowed(self, template_id: str, field_id: str) -> bool:
        """Check if Cannot Obtain is allowed for a field."""
        field = self.get_field(template_id, field_id)
        if field:
            return field.get("cno_allowed", False)
        return False

    def get_cno_default(self, template_id: str, field_id: str) -> Optional[str]:
        """Get default value to use when field is marked as CNO."""
        field = self.get_field(template_id, field_id)
        if field:
            return field.get("cno_default")
        return None

    def get_cno_flag_field(self, template_id: str, field_id: str) -> Optional[str]:
        """Get the paired CNO flag field for a value field."""
        # Look for fields where cno_flag_for matches field_id
        if template_id not in self._field_index:
            self.load_schema(template_id)

        for fid, fdef in self._field_index.get(template_id, {}).items():
            if fdef.get("cno_flag_for") == field_id:
                return fid
        return None

    def evaluate_dependencies(
        self, template_id: str, field_id: str, field_values: Dict[str, str]
    ) -> Tuple[bool, bool]:
        """
        Evaluate field dependencies to determine visibility and required status.

        Args:
            template_id: "master" or "pco"
            field_id: The field to check
            field_values: Current field values dict (field_id -> value)

        Returns:
            Tuple of (should_show, is_required)
        """
        field = self.get_field(template_id, field_id)
        if not field:
            return True, False

        base_required = field.get("required", False)
        dependencies = field.get("dependencies", [])

        if not dependencies:
            return True, base_required

        should_show = True
        is_required = base_required

        for dep in dependencies:
            dep_field_id = dep.get("field_id")
            condition = dep.get("condition")
            expected_value = dep.get("value")
            action = dep.get("action")

            actual_value = field_values.get(dep_field_id)

            # Evaluate condition
            condition_met = False
            if condition == "equals":
                condition_met = actual_value == expected_value
            elif condition == "not_equals":
                condition_met = actual_value != expected_value
            elif condition == "in":
                condition_met = actual_value in expected_value if isinstance(expected_value, list) else False
            elif condition == "not_empty":
                condition_met = actual_value is not None and actual_value != ""

            # Apply action based on condition
            if action == "show":
                should_show = condition_met
            elif action == "hide":
                should_show = not condition_met
            elif action == "require":
                if condition_met:
                    is_required = True
            elif action == "show_and_require":
                should_show = condition_met
                is_required = condition_met

        return should_show, is_required


# Singleton instance
_schema_service: Optional[SchemaService] = None


def get_schema_service() -> SchemaService:
    """Get the singleton SchemaService instance."""
    global _schema_service
    if _schema_service is None:
        _schema_service = SchemaService()
    return _schema_service
