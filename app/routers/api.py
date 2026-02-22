"""
API endpoints for data operations.
"""
import logging
import os
import sys
import json
import subprocess
import tempfile
import time
import threading
from collections import defaultdict
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Request, Form, UploadFile, File, HTTPException, Body
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from starlette.responses import StreamingResponse
from pydantic import BaseModel

from app.mock_data import add_provider, delete_provider, PROVIDERS, TEAMS, get_session_by_id, update_session
from app.models import SessionType, SessionStatus, FieldProvenance, Session
from app.services.session_service import get_session_service
from app.services.export_service import get_export_service
from app.services.schema_service import get_schema_service
from app.services.wizard_service import get_wizard_service
from app.service_context import get_active_service_dir, get_active_service
from app.services.activity_service import log_activity

router = APIRouter(prefix="/api")

# Rate limiting for login endpoints
_login_attempts: Dict[str, list] = defaultdict(list)
_LOGIN_WINDOW_SECONDS = 300  # 5 minutes
_LOGIN_MAX_ATTEMPTS = 10  # max attempts per window


def _check_rate_limit(identifier: str) -> bool:
    """Check if a login attempt is allowed. Returns True if allowed."""
    now = time.time()
    # Clean old attempts
    _login_attempts[identifier] = [t for t in _login_attempts[identifier] if now - t < _LOGIN_WINDOW_SECONDS]
    if len(_login_attempts[identifier]) >= _LOGIN_MAX_ATTEMPTS:
        return False
    _login_attempts[identifier].append(now)
    return True


# ============================================================================
# Provider Endpoints
# ============================================================================

@router.get("/providers", response_class=JSONResponse)
async def list_providers():
    """Get list of active providers for autocomplete."""
    active_providers = [
        {
            "id": p["id"],
            "name": p["name"],
            "first_name": p.get("first_name", ""),
            "last_name": p.get("last_name", ""),
            "certification": p.get("certification", ""),
            "status": p.get("status", "active"),
        }
        for p in PROVIDERS
        if p.get("status") == "active"
    ]
    return JSONResponse(content={"providers": active_providers})


@router.post("/providers")
async def create_provider(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    certification: str = Form(...),
    role: str = Form(default="Paramedic")
):
    """Create a new provider.

    Returns JSON for fetch() calls (Accept: application/json)
    Returns HTML for HTMX calls (for modal close & page reload)
    """
    name = f"{first_name} {last_name}"

    # Add provider
    provider = add_provider(
        name=name,
        first_name=first_name,
        last_name=last_name,
        certification=certification,
        role=role
    )

    try:
        slug = get_active_service()
        if slug:
            log_activity(slug, "provider_added", {"name": name})
    except Exception:
        pass

    # Check if client wants JSON (fetch calls) or HTML (HTMX)
    accept_header = request.headers.get("accept", "")

    if "application/json" in accept_header:
        # JSON response for fetch() calls from wizard
        return JSONResponse(content={
            "success": True,
            "provider": {
                "id": provider["id"],
                "name": provider["name"],
                "certification": provider["certification"],
            }
        })
    else:
        # HTML response for HTMX calls - close modal and reload page
        return HTMLResponse(content="""
        <script>
            document.getElementById('modal-container').innerHTML = '';
            window.location.reload();
        </script>
        """)


@router.delete("/providers/{provider_id}")
async def delete_provider_endpoint(provider_id: str, request: Request):
    """Delete a provider by ID."""
    success = delete_provider(provider_id)

    if not success:
        raise HTTPException(status_code=404, detail="Provider not found")

    # Check if client wants JSON or HTML
    accept_header = request.headers.get("accept", "")

    if "application/json" in accept_header:
        return JSONResponse(content={"success": True, "message": "Provider deleted"})
    else:
        # HTML response for HTMX - reload page
        return HTMLResponse(content="""
        <script>
            window.location.reload();
        </script>
        """)


# ============================================================================
# Session Endpoints
# ============================================================================

@router.post("/sessions/real-call", response_class=JSONResponse)
async def create_real_call_session(
    request: Request,
    date: str = Form(...),
    time: Optional[str] = Form(None),
    primary_provider_id: Optional[str] = Form(None),
    participant_ids: Optional[str] = Form(None),  # Comma-separated list
    platoon: Optional[str] = Form(None),
    zip_file: Optional[UploadFile] = File(None),
):
    """
    Create a new Real Call session.
    Accepts a ZIP file containing CPR report CSVs.
    Session is created with status='importing'.
    """
    service = get_session_service()

    # Parse participant IDs from comma-separated string
    parsed_participant_ids = []
    if participant_ids:
        parsed_participant_ids = [pid.strip() for pid in participant_ids.split(",") if pid.strip()]

    # Handle file upload
    uploaded_file_path = None
    original_filename = None
    if zip_file and zip_file.filename:
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
            content = await zip_file.read()
            tmp.write(content)
            uploaded_file_path = tmp.name
            original_filename = zip_file.filename

    try:
        session = service.create_real_call_session(
            date=date,
            time=time,
            primary_provider_id=primary_provider_id,
            participant_ids=parsed_participant_ids,
            uploaded_file_path=uploaded_file_path,
            original_filename=original_filename,
        )

        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "session_import", {"session_id": session["id"], "type": "real_call"})
        except Exception:
            pass

        return JSONResponse(
            status_code=201,
            content={
                "success": True,
                "session_id": session["id"],
                "status": session["status"],
                "message": "Session created. Import in progress.",
            }
        )
    finally:
        # Clean up temp file (the artifact is already copied to uploads dir)
        if uploaded_file_path and os.path.exists(uploaded_file_path):
            os.unlink(uploaded_file_path)


@router.post("/sessions/simulated", response_class=JSONResponse)
async def create_simulated_session(
    request: Request,
    date: str = Form(...),
    time: Optional[str] = Form(None),
    primary_provider_id: Optional[str] = Form(None),
    participant_ids: Optional[str] = Form(None),  # Comma-separated list
    csv_file: Optional[UploadFile] = File(None),
    paste_text: Optional[str] = Form(None),
    platoon: Optional[str] = Form(None),
    # Direct metrics (optional - if already parsed)
    duration: Optional[float] = Form(None),
    compression_rate: Optional[float] = Form(None),
    compression_depth: Optional[float] = Form(None),
    correct_depth_percent: Optional[float] = Form(None),
    correct_rate_percent: Optional[float] = Form(None),
):
    """
    Create a new Simulated session.
    Accepts either a CSV file, pasted text, or direct metrics.
    """
    service = get_session_service()

    # Parse participant IDs from comma-separated string
    parsed_participant_ids = []
    if participant_ids:
        parsed_participant_ids = [pid.strip() for pid in participant_ids.split(",") if pid.strip()]

    # Handle file upload
    csv_file_path = None
    original_filename = None
    if csv_file and csv_file.filename:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
            content = await csv_file.read()
            tmp.write(content)
            csv_file_path = tmp.name
            original_filename = csv_file.filename

    # Build metrics dict if provided directly
    metrics = None
    if any([duration, compression_rate, compression_depth, correct_depth_percent, correct_rate_percent]):
        metrics = {}
        if duration is not None:
            metrics["duration"] = duration
        if compression_rate is not None:
            metrics["compression_rate"] = compression_rate
        if compression_depth is not None:
            metrics["compression_depth"] = compression_depth
        if correct_depth_percent is not None:
            metrics["correct_depth_percent"] = correct_depth_percent
        if correct_rate_percent is not None:
            metrics["correct_rate_percent"] = correct_rate_percent

    try:
        session = service.create_simulated_session(
            date=date,
            time=time,
            primary_provider_id=primary_provider_id,
            participant_ids=parsed_participant_ids,
            metrics=metrics,
            csv_file_path=csv_file_path,
            original_filename=original_filename,
            paste_text=paste_text,
        )

        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "session_import", {"session_id": session["id"], "type": "simulated"})
        except Exception:
            pass

        return JSONResponse(
            status_code=201,
            content={
                "success": True,
                "session_id": session["id"],
                "status": session["status"],
                "message": "Session created successfully." if session["status"] == SessionStatus.COMPLETE.value else "Session created. Processing in progress.",
            }
        )
    finally:
        # Clean up temp file
        if csv_file_path and os.path.exists(csv_file_path):
            os.unlink(csv_file_path)


@router.get("/sessions/{session_id}", response_class=JSONResponse)
async def get_session(session_id: str):
    """Get session details by ID."""
    session = get_session_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    return JSONResponse(content=session)


@router.get("/sessions/{session_id}/status", response_class=JSONResponse)
async def get_session_status(session_id: str):
    """Get session status (useful for polling during import)."""
    session = get_session_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    return JSONResponse(content={
        "session_id": session["id"],
        "status": session.get("status", "unknown"),
        "error_message": session.get("error_message"),
    })


@router.post("/sessions/{session_id}/retry", response_class=JSONResponse)
async def retry_session(session_id: str):
    """
    Retry a failed session import.
    Uses the stored artifact to re-run the import.
    """
    service = get_session_service()

    can_retry, reason = service.can_retry_session(session_id)
    if not can_retry:
        raise HTTPException(status_code=400, detail=reason)

    success, message, session = service.retry_session(session_id)
    if not success:
        raise HTTPException(status_code=400, detail=message)

    # Note: Actual re-import will be triggered by ingestion service in Patch 2
    # For now, we just reset the status to 'importing'

    return JSONResponse(content={
        "success": True,
        "session_id": session["id"],
        "status": session["status"],
        "message": message,
    })


@router.get("/sessions/failed/list", response_class=JSONResponse)
async def list_failed_sessions():
    """Get all failed sessions that can be retried."""
    service = get_session_service()
    failed_sessions = service.get_failed_sessions()

    return JSONResponse(content={
        "count": len(failed_sessions),
        "sessions": [
            {
                "id": s["id"],
                "date": s.get("date"),
                "session_type": s.get("session_type"),
                "error_message": s.get("error_message"),
                "provider_name": s.get("provider_name"),
                "has_artifact": bool(s.get("artifact")),
            }
            for s in failed_sessions
        ]
    })


# ============================================================================
# Export Endpoints
# ============================================================================

@router.post("/sessions/{session_id}/export/pco", response_class=JSONResponse)
async def export_session_pco(session_id: str):
    """
    Export session data to PCO template.
    Writes to the correct month tab and appends a row with PCO minute metrics.
    """
    session = get_session_by_id(session_id)
    if not session:
        return JSONResponse(content={"success": False, "message": "Session not found"}, status_code=200)

    export_service = get_export_service()
    success, message, output_path = export_service.export_pco(session)

    if success:
        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "export", {"format": "canroc_pco", "session_id": session_id})
        except Exception:
            pass

    return JSONResponse(content={
        "success": success,
        "message": message,
        "output_file": output_path.name if output_path else None,
        "output_path": str(output_path) if output_path else None,
    })


@router.post("/sessions/{session_id}/export/master", response_class=JSONResponse)
async def export_session_master(session_id: str):
    """
    Export session data to Master template.
    Appends a row with session summary metrics.
    """
    session = get_session_by_id(session_id)
    if not session:
        return JSONResponse(content={"success": False, "message": "Session not found"}, status_code=200)

    export_service = get_export_service()
    success, message, output_path = export_service.export_master(session)

    if success:
        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "export", {"format": "canroc_master", "session_id": session_id})
        except Exception:
            pass

    return JSONResponse(content={
        "success": success,
        "message": message,
        "output_file": output_path.name if output_path else None,
        "output_path": str(output_path) if output_path else None,
    })


@router.get("/export/templates/status", response_class=JSONResponse)
async def get_export_templates_status():
    """Check which export templates are available."""
    export_service = get_export_service()
    templates = export_service.get_available_templates()

    return JSONResponse(content={
        "pco_template_available": templates["pco"],
        "master_template_available": templates["master"],
        "pco_template_path": str(export_service.settings.canroc_pco_template_path),
        "master_template_path": str(export_service.settings.canroc_master_template_path),
    })


@router.get("/exports/download/{filename}")
async def download_export(filename: str):
    """
    Download an exported Excel file.

    Args:
        filename: The filename to download (e.g., "CanROC_PCO_AllSessions_20250118_123456.xlsx")

    Returns:
        FileResponse with the Excel file
    """
    export_service = get_export_service()
    file_path = export_service.settings.export_output_dir / filename

    # Security check - ensure the file is in the exports directory
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Export file not found")

    # Ensure the file is actually in the exports directory (prevent path traversal)
    try:
        file_path.resolve().relative_to(export_service.settings.export_output_dir.resolve())
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied")

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


# ============================================================================
# CanROC Wizard Schema Endpoints
# ============================================================================

@router.get("/schemas/{template_id}", response_class=JSONResponse)
async def get_schema(template_id: str):
    """
    Get schema definition for a template.

    Args:
        template_id: "master" or "pco"

    Returns:
        Full schema JSON with pages, fields, validation rules
    """
    schema_service = get_schema_service()
    try:
        schema = schema_service.load_schema(template_id)
        return JSONResponse(content=schema)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/schemas/{template_id}/page/{page_id}", response_class=JSONResponse)
async def get_schema_page(template_id: str, page_id: int):
    """
    Get schema for a single wizard page.

    Returns:
        Page schema with fields and dependencies
    """
    schema_service = get_schema_service()
    try:
        page = schema_service.get_page(template_id, page_id)
        if not page:
            raise HTTPException(status_code=404, detail=f"Page {page_id} not found in {template_id}")
        return JSONResponse(content=page)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/schemas/{template_id}/field/{field_id}", response_class=JSONResponse)
async def get_schema_field(template_id: str, field_id: str):
    """Get schema for a single field."""
    schema_service = get_schema_service()
    field = schema_service.get_field(template_id, field_id)
    if not field:
        raise HTTPException(status_code=404, detail=f"Field {field_id} not found in {template_id}")
    return JSONResponse(content=field)


@router.get("/schemas/validate", response_class=JSONResponse)
async def validate_schemas():
    """
    Validate all schemas against Excel templates.

    Returns:
        Validation results with any drift warnings
    """
    schema_service = get_schema_service()
    results = schema_service.validate_all_schemas()

    has_errors = any(
        any("CRITICAL" in w or "MISSING" in w for w in warnings)
        for warnings in results.values()
    )

    return JSONResponse(content={
        "valid": not has_errors,
        "results": results,
    })


# ============================================================================
# CanROC Wizard State Endpoints
# ============================================================================

class FieldValueUpdate(BaseModel):
    """Request model for field value update."""
    field_id: str
    value: Optional[str] = None
    is_cno: bool = False
    cno_reason: Optional[str] = None


@router.get("/sessions/{session_id}/canroc/{template_id}", response_class=JSONResponse)
async def get_session_canroc_state(session_id: str, template_id: str):
    """
    Get current CanROC wizard state for a session.

    Returns:
        Wizard state including all field values and page statuses
    """
    session = get_session_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Get existing wizard state
    wizard_state = None
    if template_id == "master":
        wizard_state = session.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state = session.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id. Must be 'master' or 'pco'")

    if not wizard_state:
        return JSONResponse(content={
            "initialized": False,
            "session_id": session_id,
            "template_id": template_id,
            "message": "Wizard not yet initialized. POST to /start to begin.",
        })

    wizard_service = get_wizard_service()
    summary = wizard_service.get_wizard_summary(wizard_state)
    summary["initialized"] = True

    return JSONResponse(content=summary)


@router.post("/sessions/{session_id}/canroc/{template_id}/start", response_class=JSONResponse)
async def start_wizard(session_id: str, template_id: str):
    """
    Initialize wizard state for a session.
    Pre-fills fields from ZIP autofill data.
    """
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    if template_id not in ["master", "pco"]:
        raise HTTPException(status_code=400, detail="Invalid template_id. Must be 'master' or 'pco'")

    # Convert dict to Session model for service
    session = Session(**session_dict)

    wizard_service = get_wizard_service()
    wizard_state = wizard_service.initialize_wizard(session, template_id)

    # Save wizard state to session
    if template_id == "master":
        update_data = {"canroc_wizard_master": wizard_state.model_dump()}
    else:
        update_data = {"canroc_wizard_pco": wizard_state.model_dump()}

    update_session(session_id, update_data)

    summary = wizard_service.get_wizard_summary(wizard_state)
    summary["initialized"] = True

    return JSONResponse(content=summary)


@router.post("/sessions/{session_id}/canroc/{template_id}/fields", response_class=JSONResponse)
async def upsert_field_values(
    session_id: str,
    template_id: str,
    fields: List[FieldValueUpdate] = Body(...)
):
    """
    Upsert one or more field values.

    Request body:
        [
            {"field_id": "cr_age", "value": "65"},
            {"field_id": "cr_sex", "value": null, "is_cno": true, "cno_reason": "Unknown"}
        ]

    Returns:
        Updated wizard state summary
    """
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    # Get wizard state
    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        raise HTTPException(status_code=400, detail="Wizard not initialized. POST to /start first.")

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()

    # Update each field
    for field_update in fields:
        if field_update.is_cno:
            wizard_service.mark_field_cno(
                wizard_state,
                field_update.field_id,
                field_update.cno_reason
            )
        else:
            wizard_service.upsert_field(
                wizard_state,
                field_update.field_id,
                field_update.value
            )

    # Save updated wizard state
    if template_id == "master":
        update_data = {"canroc_wizard_master": wizard_state.model_dump()}
    else:
        update_data = {"canroc_wizard_pco": wizard_state.model_dump()}

    update_session(session_id, update_data)

    summary = wizard_service.get_wizard_summary(wizard_state)
    return JSONResponse(content=summary)


@router.post("/sessions/{session_id}/canroc/{template_id}/field/{field_id}/cno", response_class=JSONResponse)
async def mark_field_cno(
    session_id: str,
    template_id: str,
    field_id: str,
    reason: Optional[str] = Form(None)
):
    """Mark a field as Cannot Obtain."""
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        raise HTTPException(status_code=400, detail="Wizard not initialized")

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()
    schema_service = get_schema_service()

    # Check if CNO is allowed
    if not schema_service.is_cno_allowed(template_id, field_id):
        raise HTTPException(status_code=400, detail=f"Cannot Obtain not allowed for field {field_id}")

    wizard_service.mark_field_cno(wizard_state, field_id, reason)

    # Save
    if template_id == "master":
        update_data = {"canroc_wizard_master": wizard_state.model_dump()}
    else:
        update_data = {"canroc_wizard_pco": wizard_state.model_dump()}

    update_session(session_id, update_data)

    return JSONResponse(content={
        "success": True,
        "field_id": field_id,
        "state": "cno",
    })


@router.delete("/sessions/{session_id}/canroc/{template_id}/field/{field_id}/cno", response_class=JSONResponse)
async def clear_field_cno(session_id: str, template_id: str, field_id: str):
    """Clear CNO status from a field."""
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        raise HTTPException(status_code=400, detail="Wizard not initialized")

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()
    wizard_service.clear_field_cno(wizard_state, field_id)

    # Save
    if template_id == "master":
        update_data = {"canroc_wizard_master": wizard_state.model_dump()}
    else:
        update_data = {"canroc_wizard_pco": wizard_state.model_dump()}

    update_session(session_id, update_data)

    return JSONResponse(content={
        "success": True,
        "field_id": field_id,
        "state": "empty",
    })


@router.post("/sessions/{session_id}/canroc/{template_id}/page/{page_id}", response_class=JSONResponse)
async def save_wizard_page(
    session_id: str,
    template_id: str,
    page_id: int,
    field_values: Dict[str, Optional[str]] = Body(...)
):
    """
    Save all fields for a wizard page.
    Validates page-level rules.
    Updates page status and overall completion.
    """
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        raise HTTPException(status_code=400, detail="Wizard not initialized")

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()
    errors = wizard_service.save_page(wizard_state, page_id, field_values)

    # Save regardless of errors (partial save)
    if template_id == "master":
        update_data = {"canroc_wizard_master": wizard_state.model_dump()}
    else:
        update_data = {"canroc_wizard_pco": wizard_state.model_dump()}

    update_session(session_id, update_data)

    summary = wizard_service.get_wizard_summary(wizard_state)
    summary["page_errors"] = errors

    return JSONResponse(content=summary)


@router.get("/sessions/{session_id}/canroc/{template_id}/status", response_class=JSONResponse)
async def get_wizard_status(session_id: str, template_id: str):
    """Get wizard completion status summary."""
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        return JSONResponse(content={
            "initialized": False,
            "status": "not_started",
            "completion_percent": 0,
        })

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()
    summary = wizard_service.get_wizard_summary(wizard_state)

    return JSONResponse(content={
        "initialized": True,
        "status": summary["status"],
        "completion_percent": summary["completion_percent"],
        "current_page": summary["current_page"],
        "total_pages": summary["total_pages"],
        "can_complete": summary["can_complete"],
        "missing_required": summary["missing_required"],
    })


@router.post("/sessions/{session_id}/canroc/{template_id}/complete", response_class=JSONResponse)
async def complete_wizard(session_id: str, template_id: str):
    """
    Mark wizard as complete and normalize all blank fields to ".".
    Validates all required fields are filled or CNO.
    Updates session completion flags.
    """
    session_dict = get_session_by_id(session_id)
    if not session_dict:
        raise HTTPException(status_code=404, detail="Session not found")

    wizard_state_dict = None
    if template_id == "master":
        wizard_state_dict = session_dict.get("canroc_wizard_master")
    elif template_id == "pco":
        wizard_state_dict = session_dict.get("canroc_wizard_pco")
    else:
        raise HTTPException(status_code=400, detail="Invalid template_id")

    if not wizard_state_dict:
        raise HTTPException(status_code=400, detail="Wizard not initialized")

    from app.models import CanrocWizardState
    wizard_state = CanrocWizardState(**wizard_state_dict)

    wizard_service = get_wizard_service()
    success, errors = wizard_service.complete_wizard(wizard_state)

    if not success:
        return JSONResponse(content={
            "success": False,
            "errors": errors,
        }, status_code=400)

    # Export to payload
    payload = wizard_service.export_to_payload(wizard_state)

    # Save wizard state and update payloads
    update_data = {}
    if template_id == "master":
        update_data["canroc_wizard_master"] = wizard_state.model_dump()
        update_data["canroc_master_payload"] = payload
        update_data["canroc_master_complete"] = True
    else:
        update_data["canroc_wizard_pco"] = wizard_state.model_dump()
        update_data["canroc_pco_payload"] = payload
        update_data["canroc_pco_complete"] = True

    update_session(session_id, update_data)

    return JSONResponse(content={
        "success": True,
        "message": f"Wizard completed for {template_id}",
        "completion_percent": 100.0,
        "payload_fields": len(payload),
    })


# ============================================================================
# Bulk CanROC Export Endpoints (All Sessions)
# ============================================================================

@router.post("/canroc/export/pco", response_class=JSONResponse)
async def export_all_pco():
    """
    Export ALL sessions to a single PCO Excel file.

    Each session gets ONE row in the Excel file.
    Rows are ordered by episode date (cr_epdt) ascending.
    Missing values are normalized to "." (period).

    Returns:
        JSON with success status and filename for download
    """
    from app.services.export_service import get_export_service
    from app.services.wizard_service import get_wizard_service
    from app.services.schema_service import get_schema_service
    from app.models import CanrocWizardState
    from app.mock_data import SESSIONS

    export_service = get_export_service()
    wizard_service = get_wizard_service()
    schema_service = get_schema_service()

    # Check template exists
    if not export_service.settings.canroc_pco_template_path.exists():
        return JSONResponse(content={
            "success": False,
            "message": "PCO template not found"
        }, status_code=400)

    # Get all real_call sessions with complete status
    sessions = [
        s for s in SESSIONS
        if s.get("session_type") == "real_call" and s.get("status") == "complete"
    ]

    if not sessions:
        return JSONResponse(content={
            "success": False,
            "message": "No sessions available for export"
        }, status_code=400)

    # Sort by date ascending (oldest first for row order)
    sessions = sorted(sessions, key=lambda s: s.get("date", ""))

    try:
        # Load template
        from openpyxl import load_workbook
        from datetime import datetime

        template_path = export_service.settings.canroc_pco_template_path
        wb = load_workbook(template_path)

        # Get schema for field order
        schema = schema_service.load_schema("pco")
        all_fields = []
        for page in schema.get("pages", []):
            for field in page.get("fields", []):
                all_fields.append(field["field_id"])

        # Use the first available month tab or create "Export"
        if "Export" in wb.sheetnames:
            ws = wb["Export"]
        else:
            # Use first month tab
            month_tabs = ["Jan ", "Feb ", "Mar ", "Apr", "May", "Jun",
                         "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            ws = None
            for tab in month_tabs:
                if tab in wb.sheetnames:
                    ws = wb[tab]
                    break
                if tab.strip() in wb.sheetnames:
                    ws = wb[tab.strip()]
                    break
            if not ws:
                ws = wb.active

        # Build header -> column index mapping from Row 1
        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header_map[str(cell.value).strip()] = col_idx

        # Find starting row (after any existing data)
        start_row = 4  # Row 4 is first data row (after headers)

        # Write each session as a row
        for row_offset, session in enumerate(sessions):
            row_num = start_row + row_offset

            # Get wizard payload if available, or use stored payload
            wizard_dict = session.get("canroc_wizard_pco")
            if wizard_dict:
                wizard_state = CanrocWizardState(**wizard_dict)
                payload = wizard_service.export_to_payload(wizard_state)
            else:
                payload = session.get("canroc_pco_payload", {})

            # Write fields to row
            for field_id in all_fields:
                if field_id in header_map:
                    col_idx = header_map[field_id]
                    value = payload.get(field_id, ".")
                    if value is None:
                        value = "."
                    ws.cell(row=row_num, column=col_idx, value=value)

            # Also write session date if cr_epdt column exists
            if "cr_epdt" in header_map and "cr_epdt" not in payload:
                ws.cell(row=row_num, column=header_map["cr_epdt"], value=session.get("date", "."))

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"CanROC_PCO_AllSessions_{timestamp}.xlsx"
        output_path = export_service.settings.export_output_dir / output_filename

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save the workbook
        wb.save(output_path)
        wb.close()

        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "export", {"format": "canroc_pco", "session_count": len(sessions)})
        except Exception:
            pass

        return JSONResponse(content={
            "success": True,
            "message": f"Exported {len(sessions)} sessions to PCO template",
            "filename": output_filename,
            "session_count": len(sessions),
        })

    except Exception as e:
        logger.error(f"PCO export failed: {e}")
        return JSONResponse(content={
            "success": False,
            "message": "Export failed. Check that session data is complete."
        }, status_code=500)


@router.post("/canroc/export/master", response_class=JSONResponse)
async def export_all_master():
    """
    Export ALL sessions to a single Master Excel file.

    Each session gets ONE row in the Excel file.
    Rows are ordered by episode date (cr_epdt) ascending.
    Missing values are normalized to "." (period).

    Returns:
        JSON with success status and filename for download
    """
    from app.services.export_service import get_export_service
    from app.services.wizard_service import get_wizard_service
    from app.services.schema_service import get_schema_service
    from app.models import CanrocWizardState
    from app.mock_data import SESSIONS

    export_service = get_export_service()
    wizard_service = get_wizard_service()
    schema_service = get_schema_service()

    # Check template exists
    if not export_service.settings.canroc_master_template_path.exists():
        return JSONResponse(content={
            "success": False,
            "message": "Master template not found"
        }, status_code=400)

    # Get all real_call sessions with complete status
    sessions = [
        s for s in SESSIONS
        if s.get("session_type") == "real_call" and s.get("status") == "complete"
    ]

    if not sessions:
        return JSONResponse(content={
            "success": False,
            "message": "No sessions available for export"
        }, status_code=400)

    # Sort by date ascending (oldest first for row order)
    sessions = sorted(sessions, key=lambda s: s.get("date", ""))

    try:
        # Load template
        from openpyxl import load_workbook
        from datetime import datetime

        template_path = export_service.settings.canroc_master_template_path
        wb = load_workbook(template_path)

        # Get schema for field order
        schema = schema_service.load_schema("master")
        all_fields = []
        for page in schema.get("pages", []):
            for field in page.get("fields", []):
                all_fields.append(field["field_id"])

        # Use "Master" sheet or active sheet
        if "Master" in wb.sheetnames:
            ws = wb["Master"]
        else:
            ws = wb.active

        # Build header -> column index mapping from Row 1
        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header_map[str(cell.value).strip()] = col_idx

        # Find starting row (after any existing data)
        start_row = 4  # Row 4 is first data row (after headers)

        # Write each session as a row
        for row_offset, session in enumerate(sessions):
            row_num = start_row + row_offset

            # Get wizard payload if available, or use stored payload
            wizard_dict = session.get("canroc_wizard_master")
            if wizard_dict:
                wizard_state = CanrocWizardState(**wizard_dict)
                payload = wizard_service.export_to_payload(wizard_state)
            else:
                payload = session.get("canroc_master_payload", {})

            # Write fields to row
            for field_id in all_fields:
                if field_id in header_map:
                    col_idx = header_map[field_id]
                    value = payload.get(field_id, ".")
                    if value is None:
                        value = "."
                    ws.cell(row=row_num, column=col_idx, value=value)

            # Also write pcofile (session ID) and cr_epdt if not in payload
            if "pcofile" in header_map and "pcofile" not in payload:
                ws.cell(row=row_num, column=header_map["pcofile"], value=session.get("id", "."))
            if "cr_epdt" in header_map and "cr_epdt" not in payload:
                ws.cell(row=row_num, column=header_map["cr_epdt"], value=session.get("date", "."))

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"CanROC_Master_AllSessions_{timestamp}.xlsx"
        output_path = export_service.settings.export_output_dir / output_filename

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save the workbook
        wb.save(output_path)
        wb.close()

        try:
            slug = get_active_service()
            if slug:
                log_activity(slug, "export", {"format": "canroc_master", "session_count": len(sessions)})
        except Exception:
            pass

        return JSONResponse(content={
            "success": True,
            "message": f"Exported {len(sessions)} sessions to Master template",
            "filename": output_filename,
            "session_count": len(sessions),
        })

    except Exception as e:
        logger.error(f"Master export failed: {e}")
        return JSONResponse(content={
            "success": False,
            "message": "Export failed. Check that session data is complete."
        }, status_code=500)


# ============================================================================
# Authentication Endpoints (Desktop App)
# ============================================================================

@router.get("/auth/services")
async def list_auth_services():
    """List available services."""
    from app.service_context import list_services
    return {"services": list_services()}


@router.post("/auth/login")
async def auth_login(request: Request):
    """Authenticate and set active service."""
    from app.service_context import set_active_service
    from app.services.auth_service import check_password
    from app.desktop_config import get_service_dir

    form = await request.form()
    slug = str(form.get("service_slug", ""))
    password = str(form.get("password", ""))
    service_name = str(form.get("service_name", slug))

    # Rate limiting
    if not _check_rate_limit(f"service:{slug}"):
        return {"success": False, "error": "Too many login attempts. Please wait a few minutes."}

    service_dir = get_service_dir(slug)
    if not service_dir.exists():
        return {"success": False, "error": "Service not found"}

    if not check_password(service_dir, password):
        return {"success": False, "error": "Incorrect password"}

    set_active_service(slug, service_name)
    return {"success": True, "redirect": "/"}


@router.post("/auth/setup")
async def auth_setup(request: Request):
    """Create a new service with password."""
    from app.service_context import create_service, set_active_service
    from app.services.auth_service import hash_password

    form = await request.form()
    service_name = str(form.get("service_name", "")).strip()
    password = str(form.get("password", ""))

    if not service_name or not password:
        return {"success": False, "error": "Service name and password are required"}

    if len(password) < 8:
        return {"success": False, "error": "Password must be at least 8 characters"}

    pw_hash = hash_password(password)
    slug = create_service(service_name, pw_hash)
    set_active_service(slug, service_name)
    return {"success": True, "redirect": "/"}


@router.post("/auth/logout")
async def auth_logout():
    """Log out - clear active service and all cached data."""
    from app.service_context import clear_active_service
    clear_active_service()
    return {"success": True, "redirect": "/landing"}


@router.post("/auth/discover-services")
async def discover_services(request: Request):
    """Discover existing service backups on a GitHub repo.
    Lists subdirectories under backups/ to find service slugs.
    """
    import requests as http_requests

    data = await request.json()
    github_token = data.get("github_token", "").strip()
    repo_owner = data.get("repo_owner", "").strip()
    repo_name = data.get("repo_name", "").strip()

    if not all([github_token, repo_owner, repo_name]):
        return JSONResponse({"error": "All fields are required"}, status_code=400)

    headers = {
        "Authorization": f"token {github_token}",
        "Accept": "application/vnd.github.v3+json",
    }

    try:
        url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/backups"
        resp = http_requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 404:
            return JSONResponse({"success": True, "services": []})
        if resp.status_code != 200:
            return JSONResponse({"error": f"GitHub API returned {resp.status_code}"}, status_code=400)

        items = resp.json()
        services = [
            {"slug": item["name"], "path": item["path"]}
            for item in items
            if item.get("type") == "dir"
        ]

        return JSONResponse({"success": True, "services": services})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/auth/import-service")
async def import_service(request: Request):
    """Import a service from a GitHub backup.
    Creates the local service directory, configures backup, and restores data.
    """
    from app.services.auth_service import hash_password
    from app.services import backup_service
    from app.service_context import create_service, set_active_service
    from app.desktop_config import get_service_dir

    data = await request.json()
    github_token = data.get("github_token", "").strip()
    repo_owner = data.get("repo_owner", "").strip()
    repo_name = data.get("repo_name", "").strip()
    service_slug = data.get("service_slug", "").strip()
    service_name = data.get("service_name", "").strip() or service_slug
    password = data.get("password", "").strip()

    if not all([github_token, repo_owner, repo_name, service_slug, password]):
        return JSONResponse({"error": "All fields are required"}, status_code=400)

    if len(password) < 4:
        return JSONResponse({"error": "Password must be at least 4 characters"}, status_code=400)

    # Create the service locally
    pw_hash = hash_password(password)
    slug = create_service(service_name, pw_hash)

    # Configure backup so the service knows its GitHub repo
    service_dir = get_service_dir(slug)
    backup_service.configure(service_dir, github_token, repo_owner, repo_name)

    # Restore latest backup
    result = backup_service.list_backups(service_dir, github_token)
    restored = False
    if result.get("success") and result.get("backups"):
        newest = result["backups"][0]
        restore_result = backup_service.restore(service_dir, github_token, newest["path"])
        restored = restore_result.get("success", False)

    # Log the user into the imported service
    set_active_service(slug, service_name)

    return JSONResponse({
        "success": True,
        "slug": slug,
        "restored": restored,
        "redirect": "/",
    })


# ============================================================================
# Update Endpoints
# ============================================================================

@router.get("/updates/check")
async def check_updates(request: Request):
    """Check for application updates."""
    from app.services import update_service
    from app.version import __version__

    repo_owner = request.query_params.get("repo_owner", "")
    repo_name = request.query_params.get("repo_name", "")

    result = update_service.check_for_update(__version__, repo_owner, repo_name)
    return JSONResponse(result)


@router.get("/updates/download-stream")
async def download_update_stream(request: Request):
    """SSE endpoint that downloads the new .exe from GitHub with progress events."""
    import requests as http_requests
    from urllib.parse import urlparse

    url = request.query_params.get("url", "")
    if not url:
        async def error_stream():
            yield f"data: {json.dumps({'error': 'No download URL provided'})}\n\n"
        return StreamingResponse(error_stream(), media_type="text/event-stream")

    # C1 FIX: Whitelist allowed download URLs — only allow GitHub releases
    parsed = urlparse(url)
    allowed_hosts = {"github.com", "objects.githubusercontent.com"}
    if parsed.hostname not in allowed_hosts:
        async def error_blocked():
            yield f"data: {json.dumps({'error': 'Download URL not from an allowed source'})}\n\n"
        return StreamingResponse(error_blocked(), media_type="text/event-stream")

    if parsed.hostname == "github.com" and "/releases/download/" not in parsed.path:
        async def error_not_release():
            yield f"data: {json.dumps({'error': 'URL must be a GitHub release download'})}\n\n"
        return StreamingResponse(error_not_release(), media_type="text/event-stream")

    # Determine download directory
    if getattr(sys, 'frozen', False):
        appdata = os.environ.get("APPDATA", os.path.expanduser("~"))
        updates_dir = os.path.join(appdata, "CPR-Tracker", "_updates")
    else:
        updates_dir = os.path.join(tempfile.gettempdir(), "CPR-Tracker_updates")
    os.makedirs(updates_dir, exist_ok=True)
    dest_path = os.path.join(updates_dir, "CPR-Tracker-new.exe")

    def generate():
        try:
            resp = http_requests.get(url, stream=True, timeout=60)
            resp.raise_for_status()
            total = int(resp.headers.get("content-length", 0))
            downloaded = 0

            with open(dest_path, "wb") as f:
                for chunk in resp.iter_content(chunk_size=65536):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total > 0:
                            progress = int((downloaded / total) * 100)
                        else:
                            progress = 0
                        yield f"data: {json.dumps({'progress': progress, 'downloaded': downloaded, 'total': total})}\n\n"

            # Verify download is complete
            actual_size = os.path.getsize(dest_path)
            if total > 0 and actual_size != total:
                os.remove(dest_path)
                yield f"data: {json.dumps({'error': f'Download incomplete: got {actual_size} bytes, expected {total}'})}\n\n"
                return

            yield f"data: {json.dumps({'complete': True, 'path': dest_path, 'size': actual_size})}\n\n"
        except Exception as e:
            # Clean up partial download
            if os.path.exists(dest_path):
                try:
                    os.remove(dest_path)
                except OSError:
                    pass
            logger.error(f"Download stream error: {e}")
            yield f"data: {json.dumps({'error': 'Download failed'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@router.post("/updates/apply")
async def apply_update(request: Request):
    """Validate downloaded .exe and prepare update scripts (does not spawn them)."""
    if not getattr(sys, 'frozen', False):
        return JSONResponse({"error": "Auto-update only works in packaged mode"}, status_code=400)

    data = await request.json()
    new_exe_path = data.get("path", "")

    if not new_exe_path or not os.path.exists(new_exe_path):
        return JSONResponse({"error": "Downloaded file not found"}, status_code=400)

    # C2 FIX: Validate path is within the expected _updates directory
    appdata = os.environ.get("APPDATA", os.path.expanduser("~"))
    expected_updates_dir = os.path.normpath(os.path.join(appdata, "CPR-Tracker", "_updates"))
    resolved_path = os.path.normpath(os.path.realpath(new_exe_path))
    if not resolved_path.startswith(expected_updates_dir + os.sep):
        return JSONResponse({"error": "Update file must be in the updates directory"}, status_code=400)

    # MZ header check + size check
    try:
        file_size = os.path.getsize(new_exe_path)
        if file_size < 1_000_000:
            os.remove(new_exe_path)
            return JSONResponse({"error": "Downloaded file too small — likely corrupt"}, status_code=400)
        with open(new_exe_path, "rb") as f:
            header = f.read(2)
        if header != b"MZ":
            os.remove(new_exe_path)
            return JSONResponse({"error": "Downloaded file is not a valid executable"}, status_code=400)
    except Exception as e:
        logger.error(f"Update file validation failed: {e}")
        return JSONResponse({"error": "Cannot validate update file"}, status_code=400)

    current_exe = sys.executable
    current_pid = os.getpid()
    updates_dir = os.path.dirname(new_exe_path)
    backup_exe = os.path.join(updates_dir, "CPR-Tracker-old.exe")
    log_path = os.path.join(updates_dir, "update.log")
    bat_path = os.path.join(updates_dir, "update.bat")
    vbs_path = os.path.join(updates_dir, "update.vbs")

    # Pure .bat script for the actual update work
    bat_script = f'''@echo off
set "EXE_PATH={current_exe}"
set "NEW_EXE={new_exe_path}"
set "BACKUP={backup_exe}"
set "LOG={log_path}"
set "OLD_PID={current_pid}"

echo [%date% %time%] Update script started >> "%LOG%"
echo [%date% %time%] Waiting for PID %OLD_PID% to exit... >> "%LOG%"

REM Wait for old process to exit (up to 30s)
set /a TRIES=0
:waitloop
tasklist /FI "PID eq %OLD_PID%" 2>nul | find /i "%OLD_PID%" >nul 2>&1
if errorlevel 1 goto done_waiting
if %TRIES% GEQ 30 goto done_waiting
timeout /t 1 /nobreak >nul
set /a TRIES+=1
goto waitloop
:done_waiting

timeout /t 2 /nobreak >nul
echo [%date% %time%] Process exited, proceeding >> "%LOG%"

REM Backup current exe
copy /Y "%EXE_PATH%" "%BACKUP%" >nul 2>&1
echo [%date% %time%] Backup created >> "%LOG%"

REM Delete old exe then move new one (avoids NTFS ADS inheritance)
del /F /Q "%EXE_PATH%" >nul 2>&1
if exist "%EXE_PATH%" (
    echo [%date% %time%] ERROR: Could not delete old exe, trying copy >> "%LOG%"
    copy /Y "%NEW_EXE%" "%EXE_PATH%" >nul 2>&1
) else (
    move /Y "%NEW_EXE%" "%EXE_PATH%" >nul 2>&1
)

if not exist "%EXE_PATH%" (
    echo [%date% %time%] ERROR: New exe not in place, rolling back >> "%LOG%"
    copy /Y "%BACKUP%" "%EXE_PATH%" >nul 2>&1
    goto cleanup
)

echo [%date% %time%] File replaced successfully >> "%LOG%"

:cleanup
timeout /t 3 /nobreak >nul
del /F /Q "%NEW_EXE%" 2>nul
del /F /Q "%BACKUP%" 2>nul
echo [%date% %time%] Cleanup done >> "%LOG%"
(goto) 2>nul & del /F /Q "%~f0"
'''

    # VBScript wrapper — runs the .bat completely invisible (no terminal window)
    vbs_script = f'CreateObject("Wscript.Shell").Run "cmd /c ""{bat_path}""", 0, False\n'

    with open(bat_path, "w", encoding="utf-8") as f:
        f.write(bat_script)
    with open(vbs_path, "w", encoding="utf-8") as f:
        f.write(vbs_script)

    return JSONResponse({"success": True, "message": "Update prepared"})


@router.post("/updates/shutdown")
async def shutdown_for_update():
    """Spawn the update script then exit. Script runs invisible, waits for us to die, swaps files."""
    if not getattr(sys, 'frozen', False):
        return JSONResponse({"error": "Not in packaged mode"}, status_code=400)

    appdata = os.environ.get("APPDATA", os.path.expanduser("~"))
    vbs_path = os.path.join(appdata, "CPR-Tracker", "_updates", "update.vbs")

    if not os.path.exists(vbs_path):
        return JSONResponse({"error": "No update prepared"}, status_code=400)

    # Spawn the invisible VBScript wrapper which launches the .bat
    subprocess.Popen(
        ["wscript", "//B", vbs_path],
        creationflags=subprocess.DETACHED_PROCESS,
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    # Hard-exit after response is sent — os._exit() is intentional here:
    # sys.exit() only kills the current thread, but pywebview's main thread
    # blocks in webview.start(). We MUST terminate the entire process so the
    # update .bat script can detect PID exit and swap the executable.
    def delayed_exit():
        import time
        time.sleep(0.5)
        os._exit(0)

    threading.Thread(target=delayed_exit, daemon=True).start()
    return JSONResponse({"success": True, "message": "Shutting down for update"})


# ============================================================================
# Backup Endpoints
# ============================================================================

@router.post("/backup/configure")
async def configure_backup(request: Request):
    """Configure GitHub backup for the active service."""
    from app.services import backup_service
    service_dir = get_active_service_dir()
    if not service_dir:
        return JSONResponse({"error": "No active service"}, status_code=401)

    data = await request.json()
    github_token = data.get("github_token", "").strip()
    repo_owner = data.get("repo_owner", "").strip()
    repo_name = data.get("repo_name", "").strip()

    if not all([github_token, repo_owner, repo_name]):
        return JSONResponse({"error": "All fields are required"}, status_code=400)

    result = backup_service.configure(service_dir, github_token, repo_owner, repo_name)
    return JSONResponse(result)


@router.post("/backup/push")
async def backup_push(request: Request):
    """Trigger a backup to GitHub."""
    from app.services import backup_service
    service_dir = get_active_service_dir()
    if not service_dir:
        return JSONResponse({"error": "No active service"}, status_code=401)

    data = await request.json()
    github_token = data.get("github_token", "").strip()
    if not github_token:
        return JSONResponse({"error": "GitHub token required"}, status_code=400)

    result = backup_service.backup_now(service_dir, github_token)
    return JSONResponse(result)


@router.post("/backup/list")
async def backup_list(request: Request):
    """List available backups."""
    from app.services import backup_service
    service_dir = get_active_service_dir()
    if not service_dir:
        return JSONResponse({"error": "No active service"}, status_code=401)

    data = await request.json()
    token = data.get("github_token", "").strip()
    if not token:
        return JSONResponse({"error": "GitHub token required"}, status_code=400)

    result = backup_service.list_backups(service_dir, token)
    return JSONResponse(result)


@router.post("/backup/restore")
async def backup_restore(request: Request):
    """Restore from a GitHub backup."""
    from app.services import backup_service
    from app.service_context import _reinitialize_data

    service_dir = get_active_service_dir()
    if not service_dir:
        return JSONResponse({"error": "No active service"}, status_code=401)

    data = await request.json()
    github_token = data.get("github_token", "").strip()
    backup_path = data.get("backup_path", "").strip()

    if not github_token or not backup_path:
        return JSONResponse({"error": "Token and backup path required"}, status_code=400)

    result = backup_service.restore(service_dir, github_token, backup_path)
    if result.get("success"):
        _reinitialize_data()
    return JSONResponse(result)


@router.get("/backup/config")
async def backup_config():
    """Get current backup configuration (masked)."""
    from app.services import backup_service
    service_dir = get_active_service_dir()
    if not service_dir:
        return JSONResponse({"error": "No active service"}, status_code=401)

    result = backup_service.get_config_display(service_dir)
    return JSONResponse(result)


# ============================================================================
# Admin Endpoints
# ============================================================================

@router.post("/admin/login")
async def admin_login(request: Request):
    """Authenticate as admin."""
    from app.services.admin_service import check_admin_password, set_admin_authenticated

    # Accept both JSON and form data
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        data = await request.json()
        username = str(data.get("username", ""))
        password = str(data.get("password", ""))
    else:
        form = await request.form()
        username = str(form.get("username", ""))
        password = str(form.get("password", ""))

    if check_admin_password(username, password):
        set_admin_authenticated(True)
        return {"success": True, "redirect": "/admin"}
    return {"success": False, "error": "Invalid admin credentials"}


@router.post("/admin/logout")
async def admin_logout():
    """Log out admin."""
    from app.services.admin_service import set_admin_authenticated
    set_admin_authenticated(False)
    return {"success": True, "redirect": "/landing"}


@router.get("/admin/services-data")
async def admin_services_data():
    """Get cross-service data for admin dashboard."""
    from app.services.admin_service import is_admin_authenticated, get_all_services_data
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    data = get_all_services_data()
    return JSONResponse(content={"services": data})


# ============================================================================
# Test Data Endpoints
# ============================================================================

@router.post("/admin/test-data/generate")
async def generate_test_data_endpoint():
    """Generate test fire department data."""
    from app.services.admin_service import is_admin_authenticated
    from app.services.test_data_service import generate_test_data
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    result = generate_test_data()
    if result.get("success"):
        return JSONResponse(result)
    return JSONResponse(result, status_code=400)


@router.delete("/admin/test-data/delete")
async def delete_test_data_endpoint():
    """Delete test fire department data."""
    from app.services.admin_service import is_admin_authenticated
    from app.services.test_data_service import delete_test_data
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    result = delete_test_data()
    if result.get("success"):
        return JSONResponse(result)
    return JSONResponse(result, status_code=400)


# ============================================================================
# Annotation Endpoints (Event Markers for Trend Charts)
# ============================================================================

@router.get("/admin/annotations")
async def list_annotations():
    """Get all event annotations."""
    from app.services.admin_service import is_admin_authenticated, load_annotations
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    return JSONResponse(content={"annotations": load_annotations()})


@router.post("/admin/annotations")
async def create_annotation(request: Request):
    """Create an event annotation."""
    from app.services.admin_service import is_admin_authenticated, add_annotation
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    data = await request.json()
    month = str(data.get("month", "")).strip()
    label = str(data.get("label", "")).strip()
    description = str(data.get("description", "")).strip()
    color = str(data.get("color", "#dc2626")).strip()
    if not month or not label:
        return JSONResponse({"error": "Month and label are required"}, status_code=400)
    entry = add_annotation(month, label, description, color)
    return JSONResponse(content={"success": True, "annotation": entry})


@router.delete("/admin/annotations/{annotation_id}")
async def remove_annotation(annotation_id: str):
    """Delete an event annotation."""
    from app.services.admin_service import is_admin_authenticated, delete_annotation
    if not is_admin_authenticated():
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    if delete_annotation(annotation_id):
        return JSONResponse(content={"success": True})
    return JSONResponse({"error": "Annotation not found"}, status_code=404)


# ============================================================================
# Sync Endpoints
# ============================================================================

@router.get("/sync/status")
async def sync_status():
    """Get current auto-sync state."""
    from app.services.sync_service import get_sync_state
    return JSONResponse(content=get_sync_state())


# ============================================================================
# Settings Endpoints
# ============================================================================

@router.get("/settings/data")
async def get_settings():
    """Get settings for the active service."""
    from app.services.settings_service import load_settings
    return JSONResponse(content=load_settings())


@router.post("/settings/save")
async def save_settings_endpoint(request: Request):
    """Save all settings for the active service."""
    from app.services.settings_service import save_settings
    data = await request.json()
    success = save_settings(data)
    return JSONResponse(content={"success": success})


@router.post("/settings/{section}")
async def save_settings_section(request: Request, section: str):
    """Save a specific settings section."""
    from app.services.settings_service import update_section
    data = await request.json()
    success = update_section(section, data)
    return JSONResponse(content={"success": success})


# ============================================================================
# Report Issue Endpoint
# ============================================================================

@router.post("/report-issue")
async def report_issue(request: Request):
    """Create a GitHub issue for bug/suggestion reporting."""
    import requests as http_requests

    # Accept both JSON and form data
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        data = await request.json()
        title = str(data.get("title", "")).strip()
        description = str(data.get("description", "")).strip()
        issue_type = str(data.get("type", data.get("issue_type", "bug"))).strip()
    else:
        form = await request.form()
        title = str(form.get("title", "")).strip()
        description = str(form.get("description", "")).strip()
        issue_type = str(form.get("issue_type", "bug")).strip()

    if not title:
        return JSONResponse({"success": False, "error": "Title is required"}, status_code=400)

    # Load backup config to get GitHub token and repo info
    from app.services import backup_service
    service_dir = get_active_service_dir()

    github_token = ""
    repo_owner = ""
    repo_name = ""

    if service_dir:
        config = backup_service.get_config_display(service_dir)
        repo_owner = config.get("repo_owner", "")
        repo_name = config.get("repo_name", "")

        # Decrypt the stored token using the machine key
        bc = backup_service.load_backup_config(service_dir)
        encrypted_token = bc.get("encrypted_token", "")
        if encrypted_token:
            try:
                github_token = backup_service._decrypt_token(encrypted_token)
            except Exception:
                pass

    if not all([github_token, repo_owner, repo_name]):
        return JSONResponse({
            "success": False,
            "error": "GitHub not configured. Configure backup settings in Admin Dashboard first."
        }, status_code=400)

    # Build issue body
    labels = ["bug"] if issue_type == "bug" else ["enhancement"]
    body = f"**Type:** {issue_type.capitalize()}\n\n{description}\n\n---\n*Submitted via JcLS Tracker*"

    try:
        resp = http_requests.post(
            f"https://api.github.com/repos/{repo_owner}/{repo_name}/issues",
            headers={
                "Authorization": f"token {github_token}",
                "Accept": "application/vnd.github.v3+json",
            },
            json={
                "title": f"[{issue_type.upper()}] {title}",
                "body": body,
                "labels": labels,
            },
            timeout=15,
        )

        if resp.status_code == 201:
            issue_data = resp.json()
            return JSONResponse({
                "success": True,
                "issue_url": issue_data.get("html_url", ""),
                "issue_number": issue_data.get("number", 0),
            })
        else:
            return JSONResponse({
                "success": False,
                "error": f"GitHub API error: {resp.status_code} - {resp.text[:200]}"
            }, status_code=500)
    except Exception as e:
        logger.error(f"Report issue failed: {e}")
        return JSONResponse({
            "success": False,
            "error": "Failed to create issue. Please check your connection and try again."
        }, status_code=500)
